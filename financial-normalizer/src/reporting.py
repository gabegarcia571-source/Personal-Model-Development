"""
Financial Normalizer — Report Export Module.

Generates formatted Excel and PDF reports from pipeline output.
This module is UI-independent and can be called from the CLI or Streamlit.

Key exports:
    generate_excel(output_path, pipeline_result, session_meta, audit_trail)
    generate_pdf(output_path, pipeline_result, session_meta, audit_trail)
    generate_comparison_excel(output_path, comparison_result)
    generate_comparison_pdf(output_path, comparison_result)
"""
from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Version string embedded in every report
# ---------------------------------------------------------------------------
PIPELINE_VERSION = "1.0.0"

# ---------------------------------------------------------------------------
# Human-readable field labels
# ---------------------------------------------------------------------------
FIELD_LABELS: Dict[str, str] = {
    "revenue": "Revenue",
    "gross_profit": "Gross Profit",
    "operating_expenses": "Operating Expenses",
    "ebit": "EBIT",
    "ebitda_margin": "EBITDA Margin (%)",
    "interest_expense": "Interest Expense",
    "total_assets": "Total Assets",
    "current_assets": "Current Assets",
    "total_liabilities": "Total Liabilities",
    "current_liabilities": "Current Liabilities",
    "equity": "Shareholders' Equity",
    "total_debt": "Total Debt",
    "operating_cash_flow": "Operating Cash Flow",
    "investing_cash_flow": "Investing Cash Flow",
    "financing_cash_flow": "Financing Cash Flow",
    "entity": "Entity / Company",
    "period": "Fiscal Period",
    "industry": "Industry",
    "ebitda": "EBITDA",
    "detect_patterns": "Pattern Detection",
    "ev": "Enterprise Value",
    "output_dir": "Output Directory",
}


def _label(field: str) -> str:
    return FIELD_LABELS.get(field, field.replace("_", " ").title())


# ==========================================================================
# Excel export
# ==========================================================================

def generate_excel(
    output_path: str,
    pipeline_result: Dict[str, Any],
    session_meta: Dict[str, Any],
    audit_trail: List[Dict[str, Any]],
    anomalies: Optional[List[Dict[str, Any]]] = None,
) -> str:
    """Generate a formatted Excel workbook from pipeline output.

    Args:
        output_path: Destination file path (.xlsx).
        pipeline_result: Return value of run_pipeline().
        session_meta: Flat dict of session fields (entity, period, industry, …).
        audit_trail: List of audit event dicts from SessionState.
        anomalies: Optional list of anomaly dicts from SessionState.

    Returns:
        Absolute path of the written file.

    Raises:
        ImportError: If openpyxl is not installed.
        OSError: If the output path is not writable.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl") from exc

    anomalies = anomalies or []
    anomaly_fields = {a.get("field", "") for a in anomalies}
    output_path = str(Path(output_path).resolve())
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    ANOMALY_FILL = PatternFill("solid", fgColor="FED7AA")
    BOLD = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CURRENCY_FMT = '#,##0.00'

    def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def _style_data_row(ws, row_idx: int, n_cols: int, alt: bool, highlight: bool = False) -> None:
        fill = ANOMALY_FILL if highlight else (ALT_FILL if alt else PatternFill())
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = LEFT

    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except (TypeError, ValueError):
                    logger.debug("Failed to measure cell width for autofit in column %s", col_letter)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    def _write_df_to_sheet(ws, df: pd.DataFrame, anomaly_col: Optional[str] = None) -> None:
        headers = list(df.columns)
        ws.append([_label(h) for h in headers])
        _style_header_row(ws, 1, len(headers))
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            alt = row_idx % 2 == 0
            flag = False
            if anomaly_col and anomaly_col in df.columns:
                flag = str(row.get(anomaly_col, "")) in anomaly_fields
            values = []
            for val in row:
                if pd.isna(val):
                    values.append("")
                else:
                    values.append(val)
            ws.append(values)
            _style_data_row(ws, row_idx, len(headers), alt, highlight=flag)

    # ------------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    entity = session_meta.get("entity") or "Unknown"
    period = session_meta.get("period") or "Unknown"
    industry = session_meta.get("industry") or "—"
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    summary_rows = [
        ("Report Generated", generated_at),
        ("Pipeline Version", PIPELINE_VERSION),
        ("Entity", entity),
        ("Fiscal Period", period),
        ("Industry", industry),
        ("Pipeline Status", "Success" if pipeline_result.get("success") else "Failed"),
        ("Anomalies Detected", str(len(anomalies))),
    ]

    metrics_report = pipeline_result.get("metrics_report") or {}
    profitability = metrics_report.get("profitability", {})
    for key, val in profitability.items():
        label = _label(key.replace("_%", ""))
        if isinstance(val, float):
            if key.endswith("_%"):
                summary_rows.append((label, f"{val:.2f}%"))
            else:
                summary_rows.append((label, f"{val:,.2f}"))

    ws_summary.append(["Metric", "Value"])
    _style_header_row(ws_summary, 1, 2)
    for row_idx, (metric, value) in enumerate(summary_rows, start=2):
        ws_summary.append([metric, value])
        _style_data_row(ws_summary, row_idx, 2, row_idx % 2 == 0)
    _autofit(ws_summary)

    # ------------------------------------------------------------------
    # Sheets per output CSV file
    # ------------------------------------------------------------------
    file_map = {
        "parsed": "Transactions",
        "classified": "Classified Accounts",
        "ebitda": "EBITDA Metrics",
        "summary": "Normalized Summary",
    }
    for key, sheet_name in file_map.items():
        fpath = pipeline_result.get("files", {}).get(key)
        if not fpath or not Path(fpath).exists():
            continue
        try:
            df = pd.read_csv(fpath)
        except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError, UnicodeDecodeError) as exc:
            logger.warning("Could not read %s for Excel export: %s", fpath, exc)
            continue
        ws = wb.create_sheet(sheet_name)
        _write_df_to_sheet(ws, df)
        _autofit(ws)

    # ------------------------------------------------------------------
    # Anomalies sheet (if any)
    # ------------------------------------------------------------------
    if anomalies:
        ws_anom = wb.create_sheet("Anomalies")
        rows = []
        for a in anomalies:
            lo, hi = a.get("expected_range", (None, None))
            rows.append({
                "Field": _label(a.get("field", "")),
                "Value": a.get("value", ""),
                "Expected Min": lo,
                "Expected Max": hi,
                "Unit": a.get("unit", ""),
                "Industry": a.get("industry", ""),
                "Message": a.get("message", ""),
            })
        df_anom = pd.DataFrame(rows)
        _write_df_to_sheet(ws_anom, df_anom)
        _autofit(ws_anom)

    # ------------------------------------------------------------------
    # Audit trail sheet (read-only formatted)
    # ------------------------------------------------------------------
    ws_audit = wb.create_sheet("Audit Trail")
    ws_audit.append(["Timestamp", "Type", "Actor", "Detail"])
    _style_header_row(ws_audit, 1, 4)
    for row_idx, entry in enumerate(audit_trail, start=2):
        ts = entry.get("timestamp", "")[:19].replace("T", " ")
        ws_audit.append([
            ts,
            entry.get("entry_type", ""),
            entry.get("actor", ""),
            entry.get("detail", ""),
        ])
        _style_data_row(ws_audit, row_idx, 4, row_idx % 2 == 0)
    ws_audit.protection.sheet = True  # mark read-only
    _autofit(ws_audit)

    wb.save(output_path)
    logger.info("Excel report written to %s (%d sheets)", output_path, len(wb.sheetnames))
    return output_path


# ==========================================================================
# PDF export
# ==========================================================================

def generate_pdf(
    output_path: str,
    pipeline_result: Dict[str, Any],
    session_meta: Dict[str, Any],
    audit_trail: List[Dict[str, Any]],
    anomalies: Optional[List[Dict[str, Any]]] = None,
) -> str:
    """Generate a formatted PDF report from pipeline output.

    Args:
        output_path: Destination file path (.pdf).
        pipeline_result: Return value of run_pipeline().
        session_meta: Flat dict of session fields.
        audit_trail: List of audit event dicts.
        anomalies: Optional list of anomaly dicts.

    Returns:
        Absolute path of the written file.

    Raises:
        ImportError: If reportlab is not installed.
        OSError: If the output path is not writable.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable,
        )
    except ImportError as exc:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab") from exc

    anomalies = anomalies or []
    output_path = str(Path(output_path).resolve())
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=20, spaceAfter=8, textColor=rl_colors.HexColor("#1F4E79"))
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=14, spaceAfter=6, textColor=rl_colors.HexColor("#2E75B6"))
    H3 = ParagraphStyle("H3", parent=styles["Heading3"], fontSize=11, spaceAfter=4)
    BODY = styles["Normal"]
    SMALL = ParagraphStyle("Small", parent=BODY, fontSize=8)
    CAPTION = ParagraphStyle("Caption", parent=BODY, fontSize=9, textColor=rl_colors.grey)

    # Table styles
    HEADER_BG = rl_colors.HexColor("#1F4E79")
    ALT_BG = rl_colors.HexColor("#EBF3FB")
    ANOMALY_BG = rl_colors.HexColor("#FED7AA")
    GRID_COLOR = rl_colors.HexColor("#AAAAAA")

    def _table_style(n_rows: int, anomaly_rows: Optional[List[int]] = None) -> TableStyle:
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, GRID_COLOR),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, ALT_BG]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for r in (anomaly_rows or []):
            style.append(("BACKGROUND", (0, r), (-1, r), ANOMALY_BG))
        return TableStyle(style)

    def _df_to_table(df: pd.DataFrame, col_widths=None, anomaly_col: Optional[str] = None) -> Table:
        headers = [_label(c) for c in df.columns]
        data = [headers]
        anomaly_rows: List[int] = []
        for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
            values = []
            for val in row:
                if pd.isna(val) if not isinstance(val, str) else False:
                    values.append("")
                else:
                    values.append(Paragraph(str(val), SMALL) if isinstance(val, str) and len(str(val)) > 30 else str(val))
            if anomaly_col and anomaly_col in df.columns:
                if str(row.get(anomaly_col, "")) in {a.get("field", "") for a in anomalies}:
                    anomaly_rows.append(row_idx)
            data.append(values)
        page_width = A4[0] - 4 * cm
        t = Table(data, colWidths=col_widths or [page_width / len(headers)] * len(headers), repeatRows=1)
        t.setStyle(_table_style(len(data), anomaly_rows))
        return t

    story = []
    entity = session_meta.get("entity") or "Unknown Entity"
    period = session_meta.get("period") or "Unknown Period"
    industry = session_meta.get("industry") or "—"
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    pipeline_ok = pipeline_result.get("success", False)

    # ------------------------------------------------------------------
    # Cover page
    # ------------------------------------------------------------------
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph("Financial Normalizer", H1))
    story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(f"<b>Entity:</b> {entity}", BODY))
    story.append(Paragraph(f"<b>Fiscal Period:</b> {period}", BODY))
    story.append(Paragraph(f"<b>Industry:</b> {industry}", BODY))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"<b>Generated:</b> {generated_at}", CAPTION))
    story.append(Paragraph(f"<b>Pipeline Version:</b> {PIPELINE_VERSION}", CAPTION))
    story.append(Paragraph(
        f"<b>Pipeline Status:</b> {'✓ Success' if pipeline_ok else '✗ Failed'}",
        CAPTION,
    ))
    story.append(PageBreak())

    # ------------------------------------------------------------------
    # Executive Summary
    # ------------------------------------------------------------------
    story.append(Paragraph("Executive Summary", H2))
    metrics_report = pipeline_result.get("metrics_report") or {}
    profitability = metrics_report.get("profitability", {})
    health = metrics_report.get("health", {})

    exec_data = [["Metric", "Value"]]
    for key, val in {**profitability, **health}.items():
        if val is None:
            continue
        label = _label(key.replace("_%", ""))
        if key.endswith("_%"):
            exec_data.append([label, f"{val:.2f}%"])
        else:
            exec_data.append([label, f"{val:,.2f}"])

    exec_data.append(["Anomalies Detected", str(len(anomalies))])
    exec_data.append(["Audit Trail Entries", str(len(audit_trail))])

    if len(exec_data) > 1:
        t_exec = Table(exec_data, colWidths=[10 * cm, 6 * cm], repeatRows=1)
        t_exec.setStyle(_table_style(len(exec_data)))
        story.append(t_exec)
    else:
        story.append(Paragraph("No metrics data available.", BODY))
    story.append(Spacer(1, 0.5 * cm))

    if anomalies:
        story.append(Paragraph(f"⚠ {len(anomalies)} anomaly/anomalies detected:", H3))
        anom_data = [["Field", "Value", "Expected Range", "Message"]]
        for a in anomalies:
            lo, hi = a.get("expected_range", (None, None))
            unit = a.get("unit", "")
            rng = f"[{lo}{unit} — {hi}{unit}]" if lo is not None else "—"
            anom_data.append([
                _label(a.get("field", "")),
                f"{a.get('value', '')}{unit}",
                rng,
                Paragraph(a.get("message", ""), SMALL),
            ])
        page_width = A4[0] - 4 * cm
        t_anom = Table(anom_data, colWidths=[3 * cm, 2.5 * cm, 3 * cm, page_width - 8.5 * cm], repeatRows=1)
        t_anom.setStyle(_table_style(len(anom_data)))
        story.append(t_anom)
    else:
        story.append(Paragraph("✓ No anomalies detected.", BODY))

    story.append(PageBreak())

    # ------------------------------------------------------------------
    # Data sections per output CSV
    # ------------------------------------------------------------------
    section_titles = {
        "parsed": "Transaction Detail",
        "classified": "Classified Accounts",
        "ebitda": "EBITDA Metrics",
        "summary": "Normalized Summary",
    }
    for key, title in section_titles.items():
        fpath = pipeline_result.get("files", {}).get(key)
        if not fpath or not Path(fpath).exists():
            continue
        try:
            df = pd.read_csv(fpath)
        except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError, UnicodeDecodeError) as exc:
            logger.warning("Could not read %s for PDF export: %s", fpath, exc)
            continue
        if df.empty:
            continue
        story.append(Paragraph(title, H2))
        story.append(Paragraph(f"{len(df)} rows", CAPTION))
        story.append(Spacer(1, 0.2 * cm))
        # Limit to first 100 rows to keep PDF manageable
        df_display = df.head(100)
        if len(df) > 100:
            story.append(Paragraph(f"(Showing first 100 of {len(df)} rows)", CAPTION))
        story.append(_df_to_table(df_display))
        story.append(Spacer(1, 0.4 * cm))
        if key != list(section_titles.keys())[-1]:
            story.append(PageBreak())

    # ------------------------------------------------------------------
    # Audit trail appendix
    # ------------------------------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Appendix: Audit Trail", H2))
    story.append(Paragraph(
        f"Complete activity log — {len(audit_trail)} entries.", CAPTION
    ))
    story.append(Spacer(1, 0.3 * cm))

    if audit_trail:
        audit_data = [["Timestamp", "Type", "Actor", "Detail"]]
        for entry in audit_trail:
            ts = entry.get("timestamp", "")[:19].replace("T", " ")
            audit_data.append([
                ts,
                entry.get("entry_type", ""),
                entry.get("actor", ""),
                Paragraph(entry.get("detail", ""), SMALL),
            ])
        page_width = A4[0] - 4 * cm
        t_audit = Table(
            audit_data,
            colWidths=[3.5 * cm, 2.5 * cm, 1.5 * cm, page_width - 7.5 * cm],
            repeatRows=1,
        )
        t_audit.setStyle(_table_style(len(audit_data)))
        story.append(t_audit)
    else:
        story.append(Paragraph("No audit entries.", BODY))

    doc.build(story)
    logger.info("PDF report written to %s", output_path)
    return output_path


# ==========================================================================
# Comparison Excel export
# ==========================================================================

def generate_comparison_excel(
    output_path: str,
    comparison_result: Dict[str, Any],
    threshold_pct: float = 10.0,
) -> str:
    """Generate an Excel comparison report from a ComparisonResult dict.

    Args:
        output_path: Destination .xlsx path.
        comparison_result: Dict from comparison.compare_filings().
        threshold_pct: Delta % above which rows are highlighted amber.

    Returns:
        Absolute path of the written file.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel export.") from exc

    output_path = str(Path(output_path).resolve())
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
    RED_FILL = PatternFill("solid", fgColor="FEE2E2")
    THIN = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center")

    def _style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def _autofit(ws):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    name_a = comparison_result.get("name_a", "Filing A")
    name_b = comparison_result.get("name_b", "Filing B")

    # Main comparison sheet
    ws = wb.create_sheet("Comparison")
    headers = ["Field", name_a, name_b, "Δ (abs)", "Δ%", "Status"]
    ws.append(headers)
    _style_header(ws, len(headers))

    rows = comparison_result.get("rows", [])
    for row_idx, row in enumerate(rows, start=2):
        status = row.get("status", "ok")
        values = [
            _label(row.get("field", "")),
            row.get("value_a", ""),
            row.get("value_b", ""),
            row.get("delta_abs", ""),
            f"{row.get('delta_pct', ''):.1f}%" if isinstance(row.get("delta_pct"), float) else "—",
            status.upper(),
        ]
        ws.append(values)
        fill = (
            RED_FILL if status == "missing" else
            AMBER_FILL if status == "threshold_exceeded" else
            ALT_FILL if row_idx % 2 == 0 else PatternFill()
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = BORDER
    _autofit(ws)

    # Missing fields sheet
    missing = comparison_result.get("missing_fields", [])
    if missing:
        ws_miss = wb.create_sheet("Missing Fields")
        ws_miss.append(["Field", "Present In", "Absent From"])
        _style_header(ws_miss, 3)
        for row_idx, m in enumerate(missing, start=2):
            ws_miss.append([_label(m.get("field", "")), m.get("present_in", ""), m.get("absent_from", "")])
            ws_miss.cell(row=row_idx, column=1).border = BORDER
        _autofit(ws_miss)

    # Metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Property", "Value"])
    _style_header(ws_meta, 2)
    meta_rows = [
        ("Filing A Name", name_a),
        ("Filing B Name", name_b),
        ("Threshold %", f"{threshold_pct:.1f}%"),
        ("Matched Fields", str(len([r for r in rows if r.get("status") != "missing"]))),
        ("Missing Fields", str(len(missing))),
        ("Threshold Exceeded", str(len([r for r in rows if r.get("status") == "threshold_exceeded"]))),
        ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws_meta.append([k, v])
        ws_meta.cell(row=i, column=1).border = BORDER
        ws_meta.cell(row=i, column=2).border = BORDER
    _autofit(ws_meta)

    wb.save(output_path)
    logger.info("Comparison Excel written to %s", output_path)
    return output_path


# ==========================================================================
# Comparison PDF export
# ==========================================================================

def generate_comparison_pdf(
    output_path: str,
    comparison_result: Dict[str, Any],
    threshold_pct: float = 10.0,
) -> str:
    """Generate a PDF comparison report from a ComparisonResult dict.

    Args:
        output_path: Destination .pdf path.
        comparison_result: Dict from comparison.compare_filings().
        threshold_pct: Delta % above which rows are highlighted amber.

    Returns:
        Absolute path of the written file.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable,
        )
    except ImportError as exc:
        raise ImportError("reportlab is required for PDF export.") from exc

    output_path = str(Path(output_path).resolve())
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, textColor=rl_colors.HexColor("#1F4E79"))
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, textColor=rl_colors.HexColor("#2E75B6"))
    BODY = styles["Normal"]
    CAPTION = ParagraphStyle("Caption", parent=BODY, fontSize=9, textColor=rl_colors.grey)
    SMALL = ParagraphStyle("Small", parent=BODY, fontSize=8)

    AMBER = rl_colors.HexColor("#FEF3C7")
    RED_BG = rl_colors.HexColor("#FEE2E2")
    ALT_BG = rl_colors.HexColor("#EBF3FB")
    HEADER_BG = rl_colors.HexColor("#1F4E79")

    name_a = comparison_result.get("name_a", "Filing A")
    name_b = comparison_result.get("name_b", "Filing B")
    rows = comparison_result.get("rows", [])
    missing = comparison_result.get("missing_fields", [])
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    story = []
    story.append(Spacer(1, 2 * cm))
    story.append(Paragraph("Filing Comparison Report", H1))
    story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"<b>Filing A:</b> {name_a}", BODY))
    story.append(Paragraph(f"<b>Filing B:</b> {name_b}", BODY))
    story.append(Paragraph(f"<b>Delta Threshold:</b> {threshold_pct:.1f}%", BODY))
    story.append(Paragraph(f"<b>Generated:</b> {generated_at}", CAPTION))
    story.append(Spacer(1, 0.5 * cm))

    # Summary stats
    exceeded = [r for r in rows if r.get("status") == "threshold_exceeded"]
    story.append(Paragraph(f"Matched fields: {len(rows)} | Missing: {len(missing)} | Threshold exceeded: {len(exceeded)}", CAPTION))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph("Side-by-Side Comparison", H2))

    page_width = A4[0] - 4 * cm
    col_widths = [4*cm, 3*cm, 3*cm, 2.5*cm, 2*cm, 2*cm]
    table_data = [["Field", name_a, name_b, "Δ (abs)", "Δ%", "Status"]]
    row_styles = []
    for row_idx, row in enumerate(rows, start=1):
        status = row.get("status", "ok")
        delta_pct = row.get("delta_pct")
        table_data.append([
            _label(row.get("field", "")),
            str(row.get("value_a", "—")),
            str(row.get("value_b", "—")),
            str(row.get("delta_abs", "—")),
            f"{delta_pct:.1f}%" if isinstance(delta_pct, float) else "—",
            status.upper(),
        ])
        if status == "missing":
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), RED_BG))
        elif status == "threshold_exceeded":
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), AMBER))

    base_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#AAAAAA")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, ALT_BG]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        *row_styles,
    ])
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(base_style)
    story.append(t)

    if missing:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Missing Fields", H2))
        miss_data = [["Field", "Present In", "Absent From"]]
        for m in missing:
            miss_data.append([_label(m.get("field", "")), m.get("present_in", ""), m.get("absent_from", "")])
        t_miss = Table(miss_data, colWidths=[5*cm, 4*cm, 4*cm], repeatRows=1)
        t_miss.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#AAAAAA")),
            ("BACKGROUND", (0, 1), (-1, -1), RED_BG),
        ]))
        story.append(t_miss)

    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        "Rows highlighted in amber exceed the configured delta threshold. "
        "Rows highlighted in red indicate a field present in one filing but missing in the other.",
        CAPTION,
    ))

    doc.build(story)
    logger.info("Comparison PDF written to %s", output_path)
    return output_path
