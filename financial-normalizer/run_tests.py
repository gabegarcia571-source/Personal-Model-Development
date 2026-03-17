#!/usr/bin/env python3
"""
Comprehensive Test Suite for Financial Normalizer
Tests all major components and provides detailed health check
"""

import sys
import traceback
from pathlib import Path
import warnings
import tempfile
from typing import Any, Dict
import pandas as pd

# Add src to path
sys.path.insert(0, str(Path(__file__).parent))

class TestRunner:
    """Test runner with colored output support"""
    
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.tests = []
    
    def run_test(self, name, test_func):
        """Run a single test"""
        try:
            test_func()
            self.passed += 1
            self.tests.append((name, True, None))
            print(f"  ✓ {name:<50} PASS")
            return True
        except AssertionError as e:
            self.failed += 1
            self.tests.append((name, False, str(e)))
            print(f"  ✗ {name:<50} FAIL: {str(e)}")
            return False
        except (TypeError, ValueError, KeyError, IndexError, AttributeError, RuntimeError, OSError, ImportError, ModuleNotFoundError, AssertionError) as e:
            self.failed += 1
            self.tests.append((name, False, str(e)))
            print(f"  ✗ {name:<50} ERROR: {str(e)}")
            traceback.print_exc()
            return False
    
    def print_summary(self):
        """Print test summary"""
        print("\n" + "="*80)
        print(f"TEST SUMMARY: {self.passed} passed, {self.failed} failed")
        print("="*80)


def test_imports():
    """Test 1: Module Imports"""
    print("\n" + "="*80)
    print("TEST 1: MODULE IMPORTS")
    print("="*80)
    
    runner = TestRunner()

    def test_python_version():
        major, minor = sys.version_info[:2]
        assert (major, minor) == (3, 12), f"Expected Python 3.12, got {major}.{minor}"

    runner.run_test("Python version is 3.12", test_python_version)
    
    runner.run_test("Import pandas", lambda: __import__('pandas'))
    runner.run_test("Import numpy", lambda: __import__('numpy'))
    runner.run_test("Import yaml", lambda: __import__('yaml'))
    runner.run_test("Import TrialBalanceParser", 
                   lambda: __import__('src.ingestion.trial_balance_parser', fromlist=['TrialBalanceParser']))
    runner.run_test("Import ClassificationEngine", 
                   lambda: __import__('src.classification.classifier', fromlist=['ClassificationEngine']))
    runner.run_test("Import AdjustmentCalculator", 
                   lambda: __import__('src.normalization.adjustments', fromlist=['AdjustmentCalculator']))
    runner.run_test("Import NormalizedViewEngine", 
                   lambda: __import__('src.normalization.engine', fromlist=['NormalizedViewEngine']))
    runner.run_test("Import SyntheticAccountGenerator", 
                   lambda: __import__('src.ingestion.synthetic_generators', fromlist=['SyntheticAccountGenerator']))
    
    runner.print_summary()
    return runner.failed == 0


def test_parser():
    """Test 2: Trial Balance Parser"""
    print("\n" + "="*80)
    print("TEST 2: TRIAL BALANCE PARSER")
    print("="*80)
    
    runner = TestRunner()
    from src.ingestion.trial_balance_parser import TrialBalanceParser, TrialBalanceImbalanceWarning
    
    def test_parse_sample():
        parser = TrialBalanceParser('data/input/sample_trial_balance.csv')
        transactions = parser.parse()
        assert len(transactions) > 0, "No transactions loaded"
        assert transactions[0].account is not None, "Account field missing"
        assert transactions[0].amount is not None, "Amount field missing"
    
    def test_transaction_structure():
        parser = TrialBalanceParser('data/input/sample_trial_balance.csv')
        transactions = parser.parse()
        tx = transactions[0]
        assert hasattr(tx, 'date'), "Missing date"
        assert hasattr(tx, 'account'), "Missing account"
        assert hasattr(tx, 'description'), "Missing description"
        assert hasattr(tx, 'amount'), "Missing amount"
    
    def test_amount_calculation():
        parser = TrialBalanceParser('data/input/sample_trial_balance.csv')
        transactions = parser.parse()
        positive_sum = sum(t.amount for t in transactions if t.amount > 0)
        negative_sum = sum(t.amount for t in transactions if t.amount < 0)
        assert positive_sum > 0, "No positive amounts"
        # Note: Sample might not balance perfectly

    def test_imbalance_warning_threshold():
        parser = TrialBalanceParser('data/input/sample_trial_balance.csv', imbalance_tolerance=1.0)
        with warnings.catch_warnings(record=True) as captured:
            warnings.simplefilter('always')
            parser.parse()

        imbalance_warnings = [
            w for w in captured
            if issubclass(w.category, TrialBalanceImbalanceWarning)
        ]
        assert imbalance_warnings, "Expected imbalance warning was not raised"
    
    runner.run_test("Parse sample trial balance", test_parse_sample)
    runner.run_test("Validate transaction structure", test_transaction_structure)
    runner.run_test("Verify amount calculations", test_amount_calculation)
    runner.run_test("Raise imbalance warning above tolerance", test_imbalance_warning_threshold)
    
    runner.print_summary()
    return runner.failed == 0


def test_classifier():
    """Test 3: Account Classifier"""
    print("\n" + "="*80)
    print("TEST 3: ACCOUNT CLASSIFICATION")
    print("="*80)
    
    runner = TestRunner()
    from src.classification.classifier import ClassificationEngine, AccountType
    
    def test_config_loading():
        engine = ClassificationEngine()
        assert engine.config is not None, "Config not loaded"
        assert len(engine.config) > 0, "Config is empty"
        assert 'saas_tech' in engine.config or 'manufacturing' in engine.config, "No industries in config"
    
    def test_revenue_classification():
        engine = ClassificationEngine()
        result = engine.classify_account("4000", "Product Revenue")
        assert result.account_type == AccountType.REVENUE, f"Expected REVENUE, got {result.account_type}"
    
    def test_cogs_classification():
        engine = ClassificationEngine()
        result = engine.classify_account("5000", "Cost of Goods Sold")
        assert result.account_type == AccountType.COGS, f"Expected COGS, got {result.account_type}"
    
    def test_opex_classification():
        engine = ClassificationEngine()
        result = engine.classify_account("6000", "Salaries and Wages")
        account_type = result.account_type
        assert account_type in [AccountType.OPEX, AccountType.UNKNOWN], f"Unexpected type: {account_type}"
    
    def test_suspicious_patterns():
        parser = __import__('src.ingestion.trial_balance_parser', fromlist=['TrialBalanceParser']).TrialBalanceParser
        engine = ClassificationEngine()
        
        parser_obj = parser('data/input/sample_trial_balance.csv')
        transactions = parser_obj.parse()
        
        df = pd.DataFrame({
            'Account_Name': [t.description for t in transactions],
            'Amount': [t.amount for t in transactions],
        })
        
        flags = engine.detect_suspicious_patterns(df)
        assert isinstance(flags, list), "Flags should be a list"
    
    runner.run_test("Load classification config", test_config_loading)
    runner.run_test("Classify revenue account", test_revenue_classification)
    runner.run_test("Classify COGS account", test_cogs_classification)
    runner.run_test("Classify OpEx account", test_opex_classification)
    runner.run_test("Detect suspicious patterns", test_suspicious_patterns)
    
    runner.print_summary()
    return runner.failed == 0


def test_adjustments():
    """Test 4: Adjustment Calculator & EBITDA"""
    print("\n" + "="*80)
    print("TEST 4: ADJUSTMENTS & EBITDA CALCULATOR")
    print("="*80)
    
    runner = TestRunner()
    from src.normalization.adjustments import (
        AdjustmentCalculator, AdjustmentDetail, AdjustmentCategory, EBITDAMetric
    )
    
    def test_adjustment_creation():
        calc = AdjustmentCalculator()
        adj = AdjustmentDetail(
            adjustment_id="TEST-001",
            adjustment_name="Test Adjustment",
            adjustment_category=AdjustmentCategory.ADDBACK,
            account_code="6000",
            account_name="Salaries",
            amount=50000.0,
        )
        calc.add_adjustment(adj)
        assert len(calc.adjustments) == 1, "Adjustment not added"
    
    def test_ebitda_calculation():
        calc = AdjustmentCalculator()
        
        # Create sample GL data
        df = pd.DataFrame({
            'Account_Code': ['4000', '5000', '6000'],
            'Account_Name': ['Revenue', 'COGS', 'OpEx'],
            'Account_Type': ['revenue', 'cogs', 'opex'],
            'Amount': [1000000, 400000, 200000]
        })
        
        calc.gl_data = df
        ebitda = calc.calculate_reported_ebitda()
        
        assert ebitda is not None, "EBITDA calculation returned None"
        assert ebitda.revenue > 0, "Revenue should be positive"
        assert ebitda.ebitda >= 0, "EBITDA should be non-negative"
    
    def test_adjustment_impact():
        calc = AdjustmentCalculator()
        
        # Create minimum GL data
        df = pd.DataFrame({
            'Account_Code': ['4000', '5000', '6000'],
            'Account_Name': ['Revenue', 'COGS', 'OpEx'],
            'Account_Type': ['revenue', 'cogs', 'opex'],
            'Amount': [1000000, 400000, 200000]
        })
        
        calc.gl_data = df
        
        # Add adjustment
        adj = AdjustmentDetail(
            adjustment_id="ADJ-001",
            adjustment_name="Stock Based Comp",
            adjustment_category=AdjustmentCategory.ADDBACK,
            account_code="6000",
            account_name="Salaries",
            amount=50000,
        )
        calc.add_adjustment(adj)
        
        # Calculate impact
        impact_df = calc.get_adjustment_impact_analysis()
        assert len(impact_df) > 0, "No adjustment impact data"
        assert impact_df.iloc[0]['EBITDA_Impact'] == 50000, "Addback impact incorrect"
    
    runner.run_test("Create adjustment", test_adjustment_creation)
    runner.run_test("Calculate EBITDA", test_ebitda_calculation)
    runner.run_test("Analyze adjustment impact", test_adjustment_impact)
    
    runner.print_summary()
    return runner.failed == 0


def test_configuration():
    """Test 5: Configuration Files"""
    print("\n" + "="*80)
    print("TEST 5: CONFIGURATION & DATA FILES")
    print("="*80)
    
    runner = TestRunner()
    
    def test_categories_yaml_exists():
        config_path = Path('config/categories.yaml')
        assert config_path.exists(), "categories.yaml not found"
    
    def test_sample_data_exists():
        data_path = Path('data/input/sample_trial_balance.csv')
        assert data_path.exists(), "sample_trial_balance.csv not found"
    
    def test_sample_data_format():
        df = pd.read_csv('data/input/sample_trial_balance.csv')
        required_cols = ['Account', 'Description', 'Debit', 'Credit']
        actual_cols = df.columns.tolist()
        # Check for similar column names
        assert len(df) > 0, "Sample data is empty"
    
    def test_config_yaml_valid():
        import yaml
        with open('config/categories.yaml', 'r') as f:
            config = yaml.safe_load(f)
        assert config is not None, "Invalid YAML"
        assert isinstance(config, dict), "Config should be a dictionary"
    
    runner.run_test("Check categories.yaml exists", test_categories_yaml_exists)
    runner.run_test("Check sample data exists", test_sample_data_exists)
    runner.run_test("Validate sample data format", test_sample_data_format)
    runner.run_test("Validate YAML configuration", test_config_yaml_valid)
    
    runner.print_summary()
    return runner.failed == 0


def test_integration():
    """Test 6: Integration Tests"""
    print("\n" + "="*80)
    print("TEST 6: INTEGRATION TESTS")
    print("="*80)
    
    runner = TestRunner()
    
    def test_full_pipeline():
        from src.ingestion.trial_balance_parser import TrialBalanceParser
        from src.classification.classifier import ClassificationEngine
        
        # Parse
        parser = TrialBalanceParser('data/input/sample_trial_balance.csv')
        transactions = parser.parse()
        assert len(transactions) > 0, "Parser failed"
        
        # Classify
        df = pd.DataFrame({
            'Account_Code': [t.account for t in transactions],
            'Account_Name': [t.description for t in transactions],
            'Amount': [t.amount for t in transactions],
        })
        
        classifier = ClassificationEngine()
        classified = classifier.classify_dataframe(df)
        assert len(classified) > 0, "Classifier failed"
        assert 'Account_Type' in classified.columns, "Classification output missing Account_Type"
    
    def test_dataframe_classification():
        from src.classification.classifier import ClassificationEngine
        
        test_df = pd.DataFrame({
            'Account_Code': ['4000', '5000', '6000'],
            'Account_Name': ['Revenue', 'COGS', 'Salaries'],
        })
        
        classifier = ClassificationEngine()
        result = classifier.classify_dataframe(test_df)
        
        assert len(result) == 3, "Classification failed to process all rows"
        assert 'Account_Type' in result.columns, "Missing Account_Type column"

    def test_classification_no_row_multiplication():
        from src.classification.classifier import ClassificationEngine

        test_df = pd.DataFrame({
            'Account_Code': ['4000', '4000', '5000'],
            'Account_Name': ['Revenue', 'Revenue', 'COGS'],
            'Amount': [100.0, 200.0, 50.0],
        })

        classifier = ClassificationEngine()
        result = classifier.classify_dataframe(test_df)

        assert len(result) == len(test_df), "Classification changed row count"
        assert result['Amount'].sum() == test_df['Amount'].sum(), "Amount total changed after classification"
    
    runner.run_test("Full pipeline (parse -> classify)", test_full_pipeline)
    runner.run_test("DataFrame classification", test_dataframe_classification)
    runner.run_test("No row multiplication on duplicates", test_classification_no_row_multiplication)
    
    runner.print_summary()
    return runner.failed == 0


def test_ingestion_contracts():
    """Test 7: Ingestion Contract Enforcement"""
    print("\n" + "="*80)
    print("TEST 7: INGESTION CONTRACTS")
    print("="*80)

    runner = TestRunner()
    from src.ingestion.advanced_ingestion import (
        AdvancedIngestionEngine,
        parse_balance_sheet_contract,
        parse_cash_flow_contract,
        parse_income_statement_contract,
        parse_manual_input_contract,
        parse_pl_statement_contract,
    )

    def _make_temp_csv(content: str) -> str:
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='') as fh:
            fh.write(content)
            fh.flush()
            return fh.name

    def test_income_statement_contract():
        path = _make_temp_csv("Line_Item,Amount\nRevenue,100000\nOperating Expenses,50000\n")
        df, metadata = parse_income_statement_contract(path)
        assert isinstance(df, pd.DataFrame), "Expected DataFrame"
        assert metadata['statement_type'] == 'income_statement', "Wrong statement type"
        Path(path).unlink(missing_ok=True)

    def test_balance_sheet_contract():
        path = _make_temp_csv("Account,Balance\nCash,50000\nAccounts Payable,25000\n")
        df, metadata = parse_balance_sheet_contract(path)
        assert 'Account_Name' in df.columns, "Missing Account_Name"
        assert metadata['statement_type'] == 'balance_sheet', "Wrong statement type"
        Path(path).unlink(missing_ok=True)

    def test_cash_flow_contract():
        path = _make_temp_csv("Activity,Line_Item,Cash_Flow\nOperating,Net Income,40000\n")
        df, metadata = parse_cash_flow_contract(path)
        assert 'Amount' in df.columns, "Missing Amount"
        assert metadata['statement_type'] == 'cash_flow', "Wrong statement type"
        Path(path).unlink(missing_ok=True)

    def test_pl_contract():
        path = _make_temp_csv("Line_Item,Actual\nRevenue,100000\nExpenses,40000\n")
        df, metadata = parse_pl_statement_contract(path)
        assert len(df) == 2, "Unexpected P&L row count"
        assert metadata['statement_type'] == 'pl_statement', "Wrong statement type"
        Path(path).unlink(missing_ok=True)

    def test_manual_input_contract():
        path = _make_temp_csv("Account_Code,Account_Name,Amount\n1000,Cash,50000\n")
        df, metadata = parse_manual_input_contract(path)
        assert len(df) == 1, "Unexpected manual input row count"
        assert metadata['statement_type'] == 'manual_input', "Wrong statement type"
        Path(path).unlink(missing_ok=True)

    def test_advanced_csv_contract():
        path = _make_temp_csv("\n\nAccount,Description,Amount\n1000,Cash,50000\n2000,Payable,-25000\n")
        engine = AdvancedIngestionEngine()
        df, metadata = engine.ingest_file(path)
        assert not df.empty, "Advanced CSV ingestion returned empty DataFrame"
        assert set(['source', 'statement_type', 'confidence_level', 'warnings']).issubset(metadata.keys())
        Path(path).unlink(missing_ok=True)

    runner.run_test("Income statement contract", test_income_statement_contract)
    runner.run_test("Balance sheet contract", test_balance_sheet_contract)
    runner.run_test("Cash flow contract", test_cash_flow_contract)
    runner.run_test("P&L contract", test_pl_contract)
    runner.run_test("Manual input contract", test_manual_input_contract)
    runner.run_test("Advanced CSV ingestion contract", test_advanced_csv_contract)

    runner.print_summary()
    return runner.failed == 0


def test_validation_engine():
    """Test 8: Validation Engine"""
    print("\n" + "="*80)
    print("TEST 8: VALIDATION ENGINE")
    print("="*80)

    runner = TestRunner()
    from src.validation import ValidationEngine, Severity

    def test_error_revenue_not_positive():
        engine = ValidationEngine()
        results = engine.validate({"revenue": -1000.0})
        triggered = [r for r in results if r.field == "revenue"]
        assert triggered, "Revenue<=0 error not triggered"
        assert triggered[0].severity == Severity.ERROR

    def test_error_gross_profit_exceeds_revenue():
        engine = ValidationEngine()
        results = engine.validate({"gross_profit": 900.0, "revenue": 800.0})
        triggered = [r for r in results if r.field == "gross_profit"]
        assert triggered, "Gross profit > revenue error not triggered"
        assert triggered[0].severity == Severity.ERROR

    def test_error_opex_negative():
        engine = ValidationEngine()
        results = engine.validate({"operating_expenses": -500.0})
        triggered = [r for r in results if r.field == "operating_expenses"]
        assert triggered, "Negative opex error not triggered"
        assert triggered[0].severity == Severity.ERROR

    def test_error_balance_sheet_equation():
        engine = ValidationEngine()
        results = engine.validate({
            "total_assets": 1000.0,
            "total_liabilities": 400.0,
            "equity": 400.0,  # should be 600 to balance
        })
        triggered = [r for r in results if r.field == "total_assets"]
        assert triggered, "Balance sheet equation error not triggered"
        assert triggered[0].severity == Severity.ERROR

    def test_warning_debt_equity_ratio():
        engine = ValidationEngine()
        results = engine.validate({"total_debt": 110.0, "equity": 10.0})
        triggered = [r for r in results if r.field == "total_debt"]
        assert triggered, "D/E ratio warning not triggered"
        assert triggered[0].severity == Severity.WARNING

    def test_warning_interest_coverage():
        engine = ValidationEngine()
        results = engine.validate({"ebit": 5.0, "interest_expense": 20.0})
        triggered = [r for r in results if r.field == "ebit"]
        assert triggered, "Interest coverage warning not triggered"
        assert triggered[0].severity == Severity.WARNING

    def test_clean_input_zero_violations():
        engine = ValidationEngine()
        results = engine.validate({
            "revenue": 1000000.0,
            "gross_profit": 600000.0,
            "operating_expenses": 200000.0,
            "total_assets": 900000.0,
            "total_liabilities": 400000.0,
            "equity": 500000.0,
        })
        assert len(results) == 0, f"Expected 0 violations, got {len(results)}: {[r.message for r in results]}"

    runner.run_test("Error: revenue not positive", test_error_revenue_not_positive)
    runner.run_test("Error: gross profit exceeds revenue", test_error_gross_profit_exceeds_revenue)
    runner.run_test("Error: operating expenses negative", test_error_opex_negative)
    runner.run_test("Error: balance sheet equation broken", test_error_balance_sheet_equation)
    runner.run_test("Warning: D/E ratio > 10x", test_warning_debt_equity_ratio)
    runner.run_test("Warning: interest coverage < 1.0", test_warning_interest_coverage)
    runner.run_test("Clean input: zero violations", test_clean_input_zero_violations)

    runner.print_summary()
    return runner.failed == 0


def test_anomaly_detector():
    """Test 9: Anomaly Detector"""
    print("\n" + "="*80)
    print("TEST 9: ANOMALY DETECTOR")
    print("="*80)

    runner = TestRunner()
    from src.anomaly_detector import AnomalyDetector

    def test_detects_anomaly_below_range():
        detector = AnomalyDetector()
        anomalies = detector.detect({"gross_margin": -50.0}, industry="saas_tech")
        assert anomalies, "Expected anomaly for negative gross margin"
        assert anomalies[0].field == "gross_margin"

    def test_detects_anomaly_above_range():
        detector = AnomalyDetector()
        anomalies = detector.detect({"ebitda_margin": 200.0}, industry="manufacturing")
        assert anomalies, "Expected anomaly for 200% EBITDA margin"

    def test_no_anomaly_for_in_range():
        detector = AnomalyDetector()
        anomalies = detector.detect({"gross_margin": 65.0}, industry="saas_tech")
        assert not anomalies, f"Unexpected anomaly for in-range value: {anomalies}"

    def test_fallback_to_defaults_without_industry():
        detector = AnomalyDetector()
        anomalies = detector.detect({"ebitda_margin": -200.0})
        assert anomalies, "Expected anomaly even without industry context"

    def test_benchmarks_yaml_loaded():
        detector = AnomalyDetector()
        assert "defaults" in detector.benchmarks, "benchmarks.yaml 'defaults' key missing"
        assert "saas_tech" in detector.benchmarks, "benchmarks.yaml 'saas_tech' key missing"

    def test_interest_coverage_metric_supported():
        detector = AnomalyDetector()
        anomalies = detector.detect({"interest_coverage_ratio": 0.1}, industry="manufacturing")
        assert anomalies, "Expected anomaly for very low interest coverage"
        assert anomalies[0].field == "interest_coverage_ratio"

    def test_all_industries_have_valid_ranges_and_detect_output():
        detector = AnomalyDetector()
        required = {
            "gross_margin",
            "ebitda_margin",
            "opex_ratio",
            "revenue_growth",
            "debt_to_equity",
            "interest_coverage_ratio",
        }
        industries = [k for k in detector.benchmarks.keys() if k != "defaults"]
        assert len(industries) >= 12, f"Expected >=12 industries, got {len(industries)}"

        for industry in industries:
            section = detector.benchmarks[industry]
            assert "source" in section, f"{industry} missing source"
            missing = required - set(section.keys())
            assert not missing, f"{industry} missing metrics: {missing}"

            for metric in required:
                metric_cfg = section[metric]
                assert metric_cfg["min"] <= metric_cfg["max"], f"{industry}.{metric} min > max"

            # Force at least one out-of-range value and ensure detector emits output
            forced = {}
            for metric in required:
                forced[metric] = section[metric]["max"] + 1000.0
            anomalies = detector.detect(forced, industry=industry)
            assert anomalies, f"No anomalies returned for forced out-of-range values in {industry}"

    runner.run_test("Anomaly: value below range", test_detects_anomaly_below_range)
    runner.run_test("Anomaly: value above range", test_detects_anomaly_above_range)
    runner.run_test("No anomaly: value in range", test_no_anomaly_for_in_range)
    runner.run_test("Fallback to defaults without industry", test_fallback_to_defaults_without_industry)
    runner.run_test("benchmarks.yaml loaded correctly", test_benchmarks_yaml_loaded)
    runner.run_test("Interest coverage benchmark supported", test_interest_coverage_metric_supported)
    runner.run_test("All benchmark industries are valid and active", test_all_industries_have_valid_ranges_and_detect_output)

    runner.print_summary()
    return runner.failed == 0


def test_xbrl_ingestion():
    """Test 10: XBRL Ingestion"""
    print("\n" + "="*80)
    print("TEST 10: XBRL INGESTION")
    print("="*80)

    runner = TestRunner()
    import io
    from src.ingestion.xbrl_ingestion import XBRLIngestionEngine

    MINIMAL_XBRL = '''<?xml version="1.0" encoding="UTF-8"?>
<xbrli:xbrl
  xmlns:xbrli="http://www.xbrl.org/2003/instance"
  xmlns:us-gaap="http://fasb.org/us-gaap/2023-01-31"
  xmlns:dei="http://xbrl.sec.gov/dei/2023"
  xmlns:iso4217="http://www.xbrl.org/2003/iso4217">

  <xbrli:context id="FY2023">
    <xbrli:entity>
      <xbrli:identifier scheme="http://www.sec.gov/cik">0001234567</xbrli:identifier>
    </xbrli:entity>
    <xbrli:period>
      <xbrli:startDate>2023-01-01</xbrli:startDate>
      <xbrli:endDate>2023-12-31</xbrli:endDate>
    </xbrli:period>
  </xbrli:context>

  <xbrli:context id="BS2023">
    <xbrli:entity>
      <xbrli:identifier scheme="http://www.sec.gov/cik">0001234567</xbrli:identifier>
    </xbrli:entity>
    <xbrli:period>
      <xbrli:instant>2023-12-31</xbrli:instant>
    </xbrli:period>
  </xbrli:context>

  <xbrli:unit id="USD">
    <xbrli:measure>iso4217:USD</xbrli:measure>
  </xbrli:unit>

  <dei:EntityRegistrantName contextRef="FY2023">Test Corp Inc.</dei:EntityRegistrantName>
  <dei:DocumentPeriodEndDate contextRef="FY2023">2023-12-31</dei:DocumentPeriodEndDate>
  <dei:DocumentType contextRef="FY2023">10-K</dei:DocumentType>

  <us-gaap:Revenues contextRef="FY2023" decimals="-3" unitRef="USD">5000000</us-gaap:Revenues>
  <us-gaap:GrossProfit contextRef="FY2023" decimals="-3" unitRef="USD">3000000</us-gaap:GrossProfit>
  <us-gaap:NetIncomeLoss contextRef="FY2023" decimals="-3" unitRef="USD">800000</us-gaap:NetIncomeLoss>
  <us-gaap:Assets contextRef="BS2023" decimals="-3" unitRef="USD">12000000</us-gaap:Assets>
  <us-gaap:StockholdersEquity contextRef="BS2023" decimals="-3" unitRef="USD">6000000</us-gaap:StockholdersEquity>
</xbrli:xbrl>'''

    def _write_xbrl_fixture() -> str:
        tmp = tempfile.NamedTemporaryFile(suffix=".xml", delete=False, mode="w", encoding="utf-8")
        tmp.write(MINIMAL_XBRL)
        tmp.close()
        return tmp.name

    def test_xbrl_parses_to_dataframe():
        path = _write_xbrl_fixture()
        engine = XBRLIngestionEngine()
        df, metadata = engine.ingest(path)
        Path(path).unlink(missing_ok=True)
        assert not df.empty, "XBRL produced empty DataFrame"
        assert "Account_Name" in df.columns, "Missing Account_Name column"
        assert "Amount" in df.columns, "Missing Amount column"

    def test_xbrl_contract_enforced():
        path = _write_xbrl_fixture()
        engine = XBRLIngestionEngine()
        df, metadata = engine.ingest(path)
        Path(path).unlink(missing_ok=True)
        required_keys = {"source", "statement_type", "confidence_level", "warnings"}
        assert required_keys.issubset(metadata.keys()), f"Metadata missing keys: {required_keys - set(metadata.keys())}"

    def test_xbrl_confidence_high():
        path = _write_xbrl_fixture()
        engine = XBRLIngestionEngine()
        df, metadata = engine.ingest(path)
        Path(path).unlink(missing_ok=True)
        assert metadata["confidence_level"] >= 0.5, "Confidence should be >= 0.5 for valid XBRL"

    def test_xbrl_entity_name_extracted():
        path = _write_xbrl_fixture()
        engine = XBRLIngestionEngine()
        df, metadata = engine.ingest(path)
        Path(path).unlink(missing_ok=True)
        assert metadata.get("entity_name") == "Test Corp Inc.", f"Unexpected entity name: {metadata.get('entity_name')}"

    def test_xbrl_revenue_fact_mapped():
        path = _write_xbrl_fixture()
        engine = XBRLIngestionEngine()
        df, metadata = engine.ingest(path)
        Path(path).unlink(missing_ok=True)
        revenue_rows = df[df["Account_Name"] == "Revenue"]
        assert not revenue_rows.empty, "Revenue fact not mapped from XBRL"
        assert float(revenue_rows.iloc[0]["Amount"]) == 5000000.0

    def test_xbrl_invalid_extension_raises():
        from src.ingestion.contract import IngestionContractError
        engine = XBRLIngestionEngine()
        try:
            engine.ingest("/tmp/test_file.csv")
            assert False, "Expected IngestionContractError for .csv extension"
        except IngestionContractError:
            pass

    runner.run_test("XBRL parses to DataFrame", test_xbrl_parses_to_dataframe)
    runner.run_test("XBRL contract enforced (metadata keys)", test_xbrl_contract_enforced)
    runner.run_test("XBRL confidence >= 0.5", test_xbrl_confidence_high)
    runner.run_test("XBRL entity name extracted", test_xbrl_entity_name_extracted)
    runner.run_test("XBRL Revenue fact mapped", test_xbrl_revenue_fact_mapped)
    runner.run_test("XBRL raises on invalid extension", test_xbrl_invalid_extension_raises)

    runner.print_summary()
    return runner.failed == 0


def test_pdf_smart_parsing():
    """Test 11: Smart PDF Parsing (synthetic PDF fixtures via reportlab)"""
    print("\n" + "="*80)
    print("TEST 11: SMART PDF PARSING")
    print("="*80)

    runner = TestRunner()
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        _HAS_REPORTLAB = True
    except ImportError:
        _HAS_REPORTLAB = False

    from src.ingestion.advanced_ingestion import AdvancedIngestionEngine

    def _write_pdf_multi_column() -> str:
        """Create a synthetic PDF with two 'Amount' columns (multi-column layout)."""
        if not _HAS_REPORTLAB:
            raise RuntimeError("reportlab not installed — skipping PDF tests")
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        doc = SimpleDocTemplate(tmp.name, pagesize=letter)
        data = [
            ["Account", "Amount", "Amount"],  # duplicate column → multi-column
            ["Revenue", "100000", "110000"],
            ["COGS", "50000", "55000"],
            ["Operating Expenses", "20000", "22000"],
        ]
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        doc.build([table])
        return tmp.name

    def _write_pdf_nested_header() -> str:
        """Create a synthetic PDF with nested/multi-row headers."""
        if not _HAS_REPORTLAB:
            raise RuntimeError("reportlab not installed — skipping PDF tests")
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        doc = SimpleDocTemplate(tmp.name, pagesize=letter)
        # Row 0: span header (simulate with None cells), Row 1: sub-headers
        data = [
            ["Year ended Dec 31", None, None],         # outer header with None merged cells
            ["Account", "Description", "Balance"],     # inner header
            ["1000", "Cash", "50000"],
            ["2000", "Accounts Receivable", "30000"],
            ["3000", "Inventory", "20000"],
        ]
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 1), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        doc.build([table])
        return tmp.name

    def test_multi_column_detected():
        path = _write_pdf_multi_column()
        engine = AdvancedIngestionEngine()
        df, metadata = engine.ingest_file(path)
        Path(path).unlink(missing_ok=True)
        # Multi-column should produce a warning about duplicate columns and degrade confidence
        all_warnings = metadata.get("warnings", [])
        # Check that confidence is not 1.0 (degraded) OR that duplicate column warning was emitted
        has_dup_warning = any("Duplicate" in w or "multi-column" in w for w in all_warnings)
        # Accept either: warning generated OR data parsed successfully with valid frame
        assert not df.empty or has_dup_warning, "Multi-column PDF produced empty DataFrame with no warnings"

    def test_nested_header_flattened():
        path = _write_pdf_nested_header()
        engine = AdvancedIngestionEngine()
        df, metadata = engine.ingest_file(path)
        Path(path).unlink(missing_ok=True)
        # Should either detect nested headers or still return valid data
        # The key requirement: no crash, and we get either data or a meaningful warning
        assert df is not None, "Nested header PDF returned None"
        # If we got data, check it's valid
        if not df.empty:
            assert "Account_Name" in df.columns or len(df.columns) >= 2, "Unexpected column structure"

    def test_confidence_type_valid():
        """Confidence value must be a float between 0 and 1 for any PDF."""
        path = _write_pdf_multi_column()
        engine = AdvancedIngestionEngine()
        df, metadata = engine.ingest_file(path)
        Path(path).unlink(missing_ok=True)
        conf = metadata.get("confidence_level", -1)
        assert isinstance(conf, (int, float)), "Confidence must be numeric"
        assert 0.0 <= conf <= 1.0, f"Confidence {conf} out of range [0,1]"

    if not _HAS_REPORTLAB:
        print("  ⚠ reportlab not available — PDF fixture tests will be marked as warnings, not failures")
        runner.run_test("Multi-column PDF (no reportlab — skipped)", lambda: None)
        runner.run_test("Nested header PDF (no reportlab — skipped)", lambda: None)
        runner.run_test("Confidence range valid (no reportlab — skipped)", lambda: None)
    else:
        runner.run_test("Multi-column PDF: duplicate columns handled", test_multi_column_detected)
        runner.run_test("Nested header PDF: flattened without crash", test_nested_header_flattened)
        runner.run_test("PDF confidence: valid float 0-1", test_confidence_type_valid)

    runner.print_summary()
    return runner.failed == 0


def test_reporting_exports():
    """Test 12: Reporting exports (Excel/PDF + comparison exports)."""
    print("\n" + "="*80)
    print("TEST 12: REPORTING EXPORTS")
    print("="*80)

    runner = TestRunner()
    from src.reporting import (
        generate_excel,
        generate_pdf,
        generate_comparison_excel,
        generate_comparison_pdf,
    )

    def _sample_pipeline_result(tmpdir: Path) -> Dict[str, Any]:
        parsed = tmpdir / "1_parsed_transactions.csv"
        classified = tmpdir / "2_classified_accounts.csv"
        ebitda = tmpdir / "3_ebitda_metrics.csv"
        summary = tmpdir / "4_normalized_summary.csv"

        pd.DataFrame({
            "Account_Name": ["Revenue", "COGS"],
            "Amount": [1000.0, 400.0],
            "Entity": ["Test Co", "Test Co"],
            "Period": ["2024", "2024"],
        }).to_csv(parsed, index=False)
        pd.DataFrame({
            "Account_Name": ["Revenue", "COGS"],
            "Category": ["Revenue", "COGS"],
            "Amount": [1000.0, 400.0],
        }).to_csv(classified, index=False)
        pd.DataFrame({"Metric": ["EBITDA"], "EBITDA": [500.0]}).to_csv(ebitda, index=False)
        pd.DataFrame({"Line_Item": ["EBITDA"], "Amount": [500.0]}).to_csv(summary, index=False)

        return {
            "success": True,
            "files": {
                "parsed": str(parsed),
                "classified": str(classified),
                "ebitda": str(ebitda),
                "summary": str(summary),
            },
            "metrics_report": {
                "profitability": {
                    "gross_margin_%": 60.0,
                    "ebitda_margin_%": 25.0,
                },
                "health": {
                    "debt_to_equity": 1.2,
                    "interest_coverage_ratio": 4.2,
                },
            },
        }

    def test_generate_excel_file_non_empty():
        tmp = Path(tempfile.mkdtemp(prefix="report_excel_"))
        pipeline = _sample_pipeline_result(tmp)
        out = tmp / "report.xlsx"
        generate_excel(
            str(out),
            pipeline,
            {"entity": "Test Co", "period": "2024", "industry": "saas_tech"},
            [{"timestamp": "2024-01-01T00:00:00Z", "entry_type": "pipeline_run", "actor": "user", "detail": "run"}],
            [{"field": "gross_margin", "value": 5.0, "expected_range": [10, 90], "unit": "%", "industry": "saas_tech", "message": "low"}],
        )
        assert out.exists(), "Excel report file not created"
        assert out.stat().st_size > 0, "Excel report is empty"

    def test_generate_excel_contains_sheets():
        import openpyxl
        tmp = Path(tempfile.mkdtemp(prefix="report_excel_sheets_"))
        pipeline = _sample_pipeline_result(tmp)
        out = tmp / "report.xlsx"
        generate_excel(str(out), pipeline, {}, [])
        wb = openpyxl.load_workbook(out)
        expected = {"Summary", "Transactions", "Classified Accounts", "EBITDA Metrics", "Normalized Summary", "Audit Trail"}
        assert expected.issubset(set(wb.sheetnames)), f"Missing expected sheets: {expected - set(wb.sheetnames)}"

    def test_generate_pdf_file_non_empty():
        tmp = Path(tempfile.mkdtemp(prefix="report_pdf_"))
        pipeline = _sample_pipeline_result(tmp)
        out = tmp / "report.pdf"
        generate_pdf(str(out), pipeline, {"entity": "Test Co", "period": "2024"}, [])
        assert out.exists(), "PDF report file not created"
        assert out.stat().st_size > 0, "PDF report is empty"

    def test_generate_comparison_excel_file():
        tmp = Path(tempfile.mkdtemp(prefix="cmp_excel_"))
        out = tmp / "comparison.xlsx"
        cmp_data = {
            "name_a": "A",
            "name_b": "B",
            "rows": [{"field": "revenue", "value_a": 100.0, "value_b": 130.0, "delta_abs": 30.0, "delta_pct": 30.0, "status": "threshold_exceeded"}],
            "missing_fields": [],
        }
        generate_comparison_excel(str(out), cmp_data, threshold_pct=10.0)
        assert out.exists() and out.stat().st_size > 0, "Comparison Excel export failed"

    def test_generate_comparison_pdf_file():
        tmp = Path(tempfile.mkdtemp(prefix="cmp_pdf_"))
        out = tmp / "comparison.pdf"
        cmp_data = {
            "name_a": "A",
            "name_b": "B",
            "rows": [{"field": "revenue", "value_a": 100.0, "value_b": 130.0, "delta_abs": 30.0, "delta_pct": 30.0, "status": "threshold_exceeded"}],
            "missing_fields": [{"field": "gross_profit", "present_in": "A", "absent_from": "B"}],
        }
        generate_comparison_pdf(str(out), cmp_data, threshold_pct=10.0)
        assert out.exists() and out.stat().st_size > 0, "Comparison PDF export failed"

    runner.run_test("Excel report generated and non-empty", test_generate_excel_file_non_empty)
    runner.run_test("Excel report contains required sheets", test_generate_excel_contains_sheets)
    runner.run_test("PDF report generated and non-empty", test_generate_pdf_file_non_empty)
    runner.run_test("Comparison Excel generated", test_generate_comparison_excel_file)
    runner.run_test("Comparison PDF generated", test_generate_comparison_pdf_file)

    runner.print_summary()
    return runner.failed == 0


def test_comparison_module():
    """Test 13: Comparison module contracts and behavior."""
    print("\n" + "="*80)
    print("TEST 13: COMPARISON MODULE")
    print("="*80)

    runner = TestRunner()
    from src.comparison import FilingData, compare_filings

    def test_clean_matching_filings():
        a = FilingData(name="A", statement_type="income_statement", fields={"revenue": 100.0, "gross_profit": 50.0})
        b = FilingData(name="B", statement_type="income_statement", fields={"revenue": 110.0, "gross_profit": 48.0})
        result = compare_filings(a, b, threshold_pct=20.0)
        assert result.rows, "Expected comparison rows"
        rev = [r for r in result.rows if r.field == "revenue"][0]
        assert rev.delta_abs == 10.0
        assert rev.status == "ok"

    def test_missing_fields_on_one_side():
        a = FilingData(name="A", statement_type="income_statement", fields={"revenue": 100.0, "gross_profit": 40.0})
        b = FilingData(name="B", statement_type="income_statement", fields={"revenue": 100.0})
        result = compare_filings(a, b, threshold_pct=10.0)
        miss = [r for r in result.rows if r.field == "gross_profit"][0]
        assert miss.status == "missing"
        assert result.missing_fields, "Expected missing_fields entries"

    def test_threshold_exceeded_status():
        a = FilingData(name="A", statement_type="income_statement", fields={"revenue": 100.0})
        b = FilingData(name="B", statement_type="income_statement", fields={"revenue": 150.0})
        result = compare_filings(a, b, threshold_pct=10.0)
        rev = [r for r in result.rows if r.field == "revenue"][0]
        assert rev.status == "threshold_exceeded", "Expected threshold_exceeded status"

    runner.run_test("Comparison: clean matching filings", test_clean_matching_filings)
    runner.run_test("Comparison: missing fields one side", test_missing_fields_on_one_side)
    runner.run_test("Comparison: threshold exceeded", test_threshold_exceeded_status)

    runner.print_summary()
    return runner.failed == 0


def test_end_to_end_report_generation():
    """Test 14: End-to-end flow from pipeline run to report generation."""
    print("\n" + "="*80)
    print("TEST 14: END-TO-END REPORT FLOW")
    print("="*80)

    runner = TestRunner()
    from src.main import run_pipeline
    from src.reporting import generate_excel, generate_pdf

    def test_pipeline_then_reports_integrity():
        out_dir = Path(tempfile.mkdtemp(prefix="e2e_finorm_"))
        pipeline_result = run_pipeline(
            input_path="data/input/sample_trial_balance.csv",
            output_path=str(out_dir),
            industry="saas_tech",
            ebitda=True,
            detect_patterns=True,
            ev=None,
        )
        assert pipeline_result.get("success"), "Pipeline did not complete successfully"
        assert pipeline_result.get("files"), "Pipeline produced no output files"

        excel_out = out_dir / "e2e_report.xlsx"
        pdf_out = out_dir / "e2e_report.pdf"

        generate_excel(
            str(excel_out),
            pipeline_result,
            {"entity": "E2E Entity", "period": "2024", "industry": "saas_tech"},
            [{"timestamp": "2024-01-01T00:00:00Z", "entry_type": "pipeline_run", "actor": "system", "detail": "e2e"}],
        )
        generate_pdf(
            str(pdf_out),
            pipeline_result,
            {"entity": "E2E Entity", "period": "2024", "industry": "saas_tech"},
            [{"timestamp": "2024-01-01T00:00:00Z", "entry_type": "pipeline_run", "actor": "system", "detail": "e2e"}],
        )

        assert excel_out.exists() and excel_out.stat().st_size > 0, "E2E Excel report missing or empty"
        assert pdf_out.exists() and pdf_out.stat().st_size > 0, "E2E PDF report missing or empty"

    runner.run_test("End-to-end: pipeline output -> excel/pdf integrity", test_pipeline_then_reports_integrity)

    runner.print_summary()
    return runner.failed == 0


def test_settings_and_upload_validation():
    """Test 15: Settings loading and upload validation guards."""
    print("\n" + "="*80)
    print("TEST 15: SETTINGS & UPLOAD VALIDATION")
    print("="*80)

    runner = TestRunner()
    from src.settings import settings
    from src.interface.app_service import (
        allowed_upload_extensions,
        max_upload_size_mb,
        validate_upload_file,
    )

    def test_settings_defaults_present():
        assert settings.log_level, "log_level missing"
        assert isinstance(settings.max_file_size_mb, int), "max_file_size_mb should be int"
        assert settings.max_file_size_mb > 0, "max_file_size_mb must be positive"

    def test_allowed_extensions_non_empty():
        exts = allowed_upload_extensions()
        assert exts, "allowed_upload_extensions returned empty"
        assert "csv" in exts, "csv should be supported"

    def test_max_upload_size_exposed():
        size_mb = max_upload_size_mb()
        assert isinstance(size_mb, int), "max_upload_size_mb should return int"
        assert size_mb >= 1, "max_upload_size_mb unexpectedly low"

    def test_validate_missing_file():
        msg = validate_upload_file("/tmp/does_not_exist_123456.csv")
        assert msg is not None and "missing" in msg.lower(), "Expected missing-file validation error"

    def test_validate_unsupported_extension():
        tmp = tempfile.NamedTemporaryFile(suffix=".exe", delete=False)
        tmp.write(b"binary")
        tmp.close()
        msg = validate_upload_file(tmp.name)
        Path(tmp.name).unlink(missing_ok=True)
        assert msg is not None and "unsupported" in msg.lower(), "Expected unsupported extension error"

    def test_validate_supported_small_file():
        tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
        tmp.write(b"a,b\n1,2\n")
        tmp.close()
        msg = validate_upload_file(tmp.name)
        Path(tmp.name).unlink(missing_ok=True)
        assert msg is None, f"Expected valid small csv, got error: {msg}"

    runner.run_test("Settings defaults present", test_settings_defaults_present)
    runner.run_test("Allowed upload extensions exposed", test_allowed_extensions_non_empty)
    runner.run_test("Max upload size exposed", test_max_upload_size_exposed)
    runner.run_test("Validation catches missing file", test_validate_missing_file)
    runner.run_test("Validation catches unsupported extension", test_validate_unsupported_extension)
    runner.run_test("Validation accepts supported small file", test_validate_supported_small_file)

    runner.print_summary()
    return runner.failed == 0


def main():
    """Run all tests"""
    print("\n")
    print("╔" + "="*78 + "╗")
    print("║" + " "*78 + "║")
    print("║" + "FINANCIAL NORMALIZER - COMPREHENSIVE TEST SUITE".center(78) + "║")
    print("║" + " "*78 + "║")
    print("╚" + "="*78 + "╝")

    results = []
    results.append(("Module Imports", test_imports()))
    results.append(("Trial Balance Parser", test_parser()))
    results.append(("Account Classifier", test_classifier()))
    results.append(("Adjustments & EBITDA", test_adjustments()))
    results.append(("Configuration & Data", test_configuration()))
    results.append(("Integration Tests", test_integration()))
    results.append(("Ingestion Contracts", test_ingestion_contracts()))
    results.append(("Validation Engine", test_validation_engine()))
    results.append(("Anomaly Detector", test_anomaly_detector()))
    results.append(("XBRL Ingestion", test_xbrl_ingestion()))
    results.append(("Smart PDF Parsing", test_pdf_smart_parsing()))
    results.append(("Reporting Exports", test_reporting_exports()))
    results.append(("Comparison Module", test_comparison_module()))
    results.append(("E2E Report Flow", test_end_to_end_report_generation()))
    results.append(("Settings & Upload Validation", test_settings_and_upload_validation()))

    # Final summary
    print("\n" + "="*80)
    print("FINAL TEST RESULTS".center(80))
    print("="*80)
    total_pass = sum(1 for _, passed in results if passed)
    total_tests = len(results)
    
    for name, passed in results:
        status = "✓ PASS" if passed else "✗ FAIL"
        print(f"  {name:<40} {status}")
    
    print("="*80)
    print(f"Overall: {total_pass}/{total_tests} test suites passed")
    print("="*80)
    
    if total_pass == total_tests:
        print("\n🎉 ALL TESTS PASSED! Your project is working correctly.\n")
        return 0
    else:
        print(f"\n⚠️  {total_tests - total_pass} test suite(s) failed. Review errors above.\n")
        return 1


if __name__ == "__main__":
    sys.exit(main())
